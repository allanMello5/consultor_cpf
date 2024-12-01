import openpyxl
#selenium permiti abrir o navegador e interagir com a web
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

#1-Entrar na planilha e  extrair o cpf
planilha_clientes = openpyxl.load_workbook('dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

 # 2 entrar no site e usar o cpf para pesquisar o stuatus do pagamento daquele cliente. #2entro no site 10909090e9 para pesquisar o status do pagamento  daquele cliente
driver = webdriver.Chrome()
driver.get('http://localhost:52330/cpf_test_site/cpf.html')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha  
    sleep(5)
    campo_pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(1)
    campo_pesquisa.send_keys(cpf)
    sleep(2)
    #vericar se esta em  dia ou atrasado
    botao_pesquisar = driver.find_element(By.XPATH,"//button[@class='']")
    sleep(1)
    botao_pesquisar.click()
    sleep(4)
    #4 se estiver em dia , pegar data do pagamento e o metodo de pagamento
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
    status.text
    if status.text =='em dia':
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMehod']")

        pagina_fechamento.append([nome, valor, cpf , vencimento, 'em dia', 'xxx','xxx' ])
    else:
        planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
        pagina_fechamento = planilha_fechamento['Planilha1']
        
        pagina_fechamento.append([nome, valor, cpf, vencimento,'pendente' ])